import openpyxl
import time


def dealDate(companyNameList,companyPidList):
    current_datetime = time.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = current_datetime + "_file.xlsx"
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1['A1'] = "控股/投资公司"
    sheet1['B1'] = "PID"
    nameColumn = 1
    pidColumn = 2
    for row_index,value1 in enumerate(companyNameList,start=2):
        sheet1.cell(row=row_index,column=nameColumn).value = value1
    for row_index,value2 in enumerate(companyPidList,start=2):
        sheet1.cell(row=row_index,column=pidColumn).value = value2
    print("\033[0;34m【+】\033[0m" + "数据以去重,正在保存输出结果")
    time.sleep(1)
    workbook.save(file_name)